#!/usr/bin/python
#coding=utf-8
#xinghe <xingh3223@berryoncology.com>
import xlrd
import xlwt
import openpyxl as xl
from xlutils.copy import copy
import os
from collections import OrderedDict,deque
import glob
import argparse
import subprocess

'''合并表格,多个sheet中含有多个子table的情况,可以添加列也可以添加行,可以指定title的行数.'''

def judgeVanList(alist):
	'''判断某行是否全为空.'''
	flag = False
	for each in alist:
		if each.strip() != '':
			flag = True
			break
	return flag


def reverseTable(aList):
	'''翻转嵌套List,如[[1,2,3],[4,5,6]]转换成[[1,4],[2,5],[3,6]]'''
	newContentList = []
	for j in range(len(aList[0])):
		newSubList = []
		for i in range(len(aList)):
			newSubList.append(aList[i][j])
		newContentList.append(newSubList)
	return newContentList


def getTablesOfSheet(workbook,sheetName,directions = []):
	'''一个sheet中可能有多个子table,将sheet中所有table放到字典中'''

	wb = xlrd.open_workbook(workbook)
	rs = wb.sheet_by_name(sheetName)
	lastFlag = False
	content = OrderedDict()
	part = 0 
	for i in range(0,rs.nrows):
		row = rs.row_values(i)
		thisFlag = judgeVanList(row)
		if lastFlag == False and thisFlag == True:
			part += 1
			content['part'+str(part)] = []
			content['part'+str(part)].append(row)
		else:
			if thisFlag == True:
				content['part'+str(part)].append(row)
		lastFlag = thisFlag

	if directions == []:
		directions = [ 'add_row' for i in range(len(content.keys())) ]
	if len(directions) != len(content.keys()):
		sys.exit('如果填写directions参数,那么它的个数必须与单个sheet中table数相同.')
	directions = deque(directions)
	
	for ith,dir in enumerate(directions):
		if dir == 'add_col':
			revPart = 'part'+str(ith+1)
			newRevTab = reverseTable(content[revPart])
			content[revPart] = newRevTab				
	return(content)
	

def getTitleOfTables(TableDic,titleLines,directions):
	'''获取子table的title和data部分,title可能为多行,titleLines指定每个子table的title行数.'''

	titleLines = deque(list(titleLines))
	NewTableDic = OrderedDict()
	for part,table in TableDic.items():
		line = titleLines.popleft()
		if int(line) > 0:
			title = table[0:int(line)]
			data = table[int(line):]
		elif int(line) == 0:
			title = []
			data = table[:]
		else:
			sys.exit('title的行数应该是一个正整数.')
		NewTableDic[part] = [title,data]
	return NewTableDic


def getAll(workbook,sheetName,TableTitleLines,directions):
	'''获取一个sheet中所有子table(区分title和data内容).'''
	tables = getTablesOfSheet(workbook,sheetName,directions)
	tableOfSheet = getTitleOfTables(tables,TableTitleLines,directions)
	return tableOfSheet


def MergedSheetOfExcels(workbookList,sheetName,TableTitleLines = [],directions = []):
	'''合并多个excel表中同名sheet中的对应的table.'''
	basicStat = OrderedDict()
	for workbook in workbookList:
		tablesOfSheet = getAll(workbook,sheetName,TableTitleLines,directions)
		for part,content in tablesOfSheet.items():
			title = content[0]
			data = content[1]
			if part not in basicStat.keys():
				basicStat[part] = []
				basicStat[part].extend(title)
			basicStat[part].extend(data)
	return(basicStat)


def writeExcel(MergedSheetOfExcels_Dic,sheetName,outputFileName):
	'''将合并的数据写入excel表'''

	if os.path.exists(outputFileName):
		rb = xlrd.open_workbook(outputFileName,formatting_info = True,on_demand = True)
		wb = copy(rb)
	else:		
		wb = xlwt.Workbook()
	tblDic = MergedSheetOfExcels_Dic
	ws = wb.add_sheet(sheetName)
	startRow = 0
	for part,content in tblDic.items():
		for rowNum in range(0,len(content)):
			for colNum in range(0,len(content[rowNum])):
				ws.write(rowNum+startRow,colNum,content[rowNum][colNum])
		startRow += len(content)+2
	wb.save(outputFileName)	   


def writeExcel2(MergedSheetOfExcels_Dic,sheetName,outputFileName):
	'''将合并的数据写入excel表'''

	if os.path.exists(outputFileName):
		wb = xl.load_workbook(outputFileName)
	else:		
		wb = xl.Workbook()
	tblDic = MergedSheetOfExcels_Dic
	ws = wb.create_sheet(sheetName)
	startRow = 0
	for part,content in tblDic.items():
		for rowNum in range(0,len(content)):
			for colNum in range(0,len(content[rowNum])):
				ws.cell(rowNum+startRow+1,colNum+1,content[rowNum][colNum])
		startRow += len(content)+2
	wb.save(outputFileName)	   


def getROSList(analysisDir):
	ros = subprocess.Popen('grep -l ROS %s/*/sample.cfg.*'%(analysisDir),stdin = subprocess.PIPE,stdout = subprocess.PIPE,stderr = subprocess.PIPE,universal_newlines=True,shell=True)
	RmXLSList = []
	for each in ros.stdout:
		print(each)
		subdir = os.path.dirname(each.strip())
		ROSExcel = subdir + '/' + os.path.basename(subdir) + '.xls'
		if os.path.isfile(ROSExcel):
			RmXLSList.append(ROSExcel)
	return RmXLSList


def getAllExcels(analysisDir,ROS = 'n'):
	sets = glob.glob(analysisDir+'/*')
	workbookListTmp = []
	for each in sets:
		each = each.strip()
		rootDir = os.path.basename(each)
		excel = each + '/' + rootDir+'.xls'
		if os.path.isfile(excel):
			workbookListTmp.append(excel)
	RmList = []
	if ROS.lower() == 'n':
		RmList = getROSList(analysisDir)
		workbookListTmp = list(set(workbookListTmp)-set(RmList))		
	return workbookListTmp


class info(object):
	def __init__(self,sheetName,TitleLineNum = [],Direction = []):
		self.SN = sheetName
		self.TLN = TitleLineNum
		self.Dir = Direction

if __name__ == '__main__':
	parser = argparse.ArgumentParser(prog = '',description = '合并表格')
	parser.add_argument("-o","--out", help = '输出文件名',default='MergedExcels.xlsx')
	parser.add_argument("-R","--ROS", help = 'y:包括ROS1结果;n:不包括ROS1结果;r:只有ROS1的结果',default='n')
	args = parser.parse_args()

	analysisDir = '/share/Oncology/production/cSMART/cSMART/analysis/xinghe/HangZhouHospital_KY/analysis'
	if args.ROS.lower() == 'y':
		workbookList = getAllExcels(analysisDir,ROS='y') 
	elif args.ROS.lower() == 'n':
		workbookList = getAllExcels(analysisDir,ROS='n') 
	elif args.ROS.lower() == 'r':
		workbookList = getROSList(analysisDir)
	else:
		sys.exit('-R参数值的范围: y,包括ROS1结果;n,不包括ROS1结果;r,只有ROS1的结果')

	S1 = info('QC',[1])	
	S2 = info('summary',[1,1])
	S3 = info('all.must.given',[1])
	S4 = info('mutation',[0],['add_col'])
	S5 = info('fusion',[1])
	S6 = info('fusion.detail',[1])
	S7 = info('basic stat',[2,1,0])
	S8 = info('primer stat',[0],['add_col'])
	sheetTitleList = [S1,S2,S3,S4,S5,S6,S7,S8]

	for each in sheetTitleList:
		MergedSheetOfExcels_Dic = MergedSheetOfExcels(workbookList,each.SN,each.TLN,each.Dir)
		writeExcel2(MergedSheetOfExcels_Dic,each.SN,args.out)
		

